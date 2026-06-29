import scrapy
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment


class TableSpider(scrapy.Spider):
    name = 'market'
    start_urls = ['https://www.sharesansar.com/today-share-price']

    def parse(self, response):
        # Extracting the table rows
        rows = response.xpath('//table//tr')

        # Defining a list to store the rows
        table_data = []

        # Extracting the table headers
        header_row = rows[0]
        header_cells = header_row.xpath('.//th//text()').getall()
        # Removing whitespace from the header cells
        header_cells = [cell.strip() for cell in header_cells]

        # Adding the header row to the table data list
        table_data.append(header_cells)

        # Looping through each row after the header
        for row in rows[1:]:
            # Extracting the text from each cell
            cells = row.xpath('.//td//text()').getall()
            # Removing whitespace from the cells
            cells = [cell.strip() for cell in cells]
            # Filter out empty values from the cells
            cells = [cell for cell in cells if cell]
            # If any cell in the row has data, add it to the table data list
            if any(cells):
                table_data.append(cells)

        # Guard: a header-only scrape means the page had no price rows
        # (rendering error, layout change, etc.) — never write an empty file.
        if len(table_data) <= 1:
            self.logger.warning(
                'Scraped table has no data rows; skipping write (nothing to archive).'
            )
            return

        # Guard: NEPSE is closed on weekends and public holidays, on which
        # ShareSansar keeps showing the previous trading session. That scrape is
        # byte-for-byte identical to the latest CSV we already have, so skip it
        # instead of committing a duplicate day. (No hardcoded holiday calendar
        # needed — the data itself tells us nothing new happened.)
        date_str = datetime.now().strftime('%Y_%m_%d')
        file_path = f'docs/Data/{date_str}.csv'
        if self.is_duplicate_of_latest(table_data, skip_file=f'{date_str}.csv'):
            self.logger.info(
                'Scraped data matches the latest archived session (market likely '
                'closed); skipping write.'
            )
            return

        # Converting the table_data list into a DataFrame
        df = pd.DataFrame(table_data)

        # Saving the DataFrame to a CSV file
        df.to_csv(file_path, header=False, index=False)

        # combine all csv to single excel with worksheets
        self.update_combined_excel()

    @staticmethod
    def rows_without_serial(rows):
        # Drop the leading S.No column so identical sessions compare equal even
        # if the serial numbering ever shifts.
        return [tuple(str(c) for c in row[1:]) for row in rows]

    def is_duplicate_of_latest(self, table_data, skip_file):
        """True when the freshly scraped table equals the most recent CSV."""
        data_dir = 'docs/Data'
        if not os.path.isdir(data_dir):
            return False
        existing = sorted(
            f for f in os.listdir(data_dir)
            if f.endswith('.csv') and f != skip_file
        )
        if not existing:
            return False

        latest_path = os.path.join(data_dir, existing[-1])
        try:
            prev = pd.read_csv(
                latest_path, header=None, dtype=str, keep_default_na=False
            ).values.tolist()
        except Exception as e:
            self.logger.warning(f'Could not read {latest_path} for dedupe: {e}')
            return False

        # Compare data rows only (skip the header row in each), ignoring S.No.
        return self.rows_without_serial(table_data[1:]) == self.rows_without_serial(prev[1:])

    def update_combined_excel(self):
        csv_files = [file for file in os.listdir('docs/Data') if file.endswith('.csv')]

        # sort so that latest csv is the first worksheet
        csv_files.sort(reverse=True)

        excel_file_path = 'docs/Data/combined_excel.xlsx'
        txt_file_path = 'docs/Data/list_of_csv_files.txt'


        if os.path.exists(excel_file_path): 
            # If the file exists, append only the latest csv to the workbook.
            try:
                workbook = load_workbook(excel_file_path)
                sheet_to_add = csv_files[0]
                sheet_to_add_name = csv_files[0].replace('.csv', '')
                sheetNames = workbook.sheetnames

                if sheet_to_add_name in sheetNames:
                    # Deleting the worksheet if it already exists to replace data
                    workbook.remove(workbook[sheet_to_add_name])
                
                df = pd.read_csv(f'docs/Data/{sheet_to_add}')
                new_sheet = workbook.create_sheet(title=sheet_to_add_name, index=0)
                

                for col_index, header in enumerate(df.columns, start=1):
                    new_sheet.cell(row=1, column=col_index, value=header)

                for col_index in range(1, len(df.columns) + 1):
                    cell = new_sheet.cell(row=1, column=col_index)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center') 


                border_style = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='medium'), 
                        bottom=Side(style='thin')) 

                for col_index in range(1, len(df.columns) + 1):
                    cell = new_sheet.cell(row=1, column=col_index)
                    cell.border = border_style
                
                for row_index, row in df.iterrows():
                    for col_index, value in enumerate(row, start=1):
                        new_sheet.cell(row=row_index + 2, column=col_index, value=value)
        
                workbook.save(excel_file_path)

            except Exception as e:
                print(f"Error: {e}")

        else: 
            # If file doesn't already exist, take all csv and write new excel file.
            with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                for csv_file in csv_files:
                    df = pd.read_csv(f'docs/Data/{csv_file}')
                    sheet_name = os.path.splitext(csv_file)[0]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Update or create the list_of_csv_files.txt
        with open(txt_file_path, 'w') as file:
            for csv_file in csv_files:
                file.write(csv_file + '\n')